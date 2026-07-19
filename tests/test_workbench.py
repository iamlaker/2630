import tempfile
import time
import unittest
from pathlib import Path

from forecast_engine import InMemoryWorkbookEngine, WarmExcelWorkerError, run_forward_calculation
from openpyxl import load_workbook
from workbench import WorkbenchService, apply_linkage, filter_parameters
from reverse_calculation import search_priority_variables, search_single_variable


YEARS = tuple(range(2026, 2031))


class ProductionWorkbenchUiTests(unittest.TestCase):
    def setUp(self):
        self.web = Path(__file__).parents[1] / "web"

    def test_production_entry_uses_unified_four_module_shell(self):
        html = (self.web / "index.html").read_text(encoding="utf-8")
        for module in ("forward", "single", "multi", "rules"):
            self.assertIn(f'data-module="{module}"', html)
        self.assertIn('id="workspace"', html)
        self.assertEqual(html.count('class="pane-resizer"'), 2)
        self.assertIn('id="mobilePaneTabs"', html)
        self.assertIn('/workbench-state.js', html)
        self.assertNotIn('workbench-prototype', html)

    def test_state_model_is_persistent_and_module_scoped(self):
        source = (self.web / "workbench-state.js").read_text(encoding="utf-8")
        self.assertIn('const MODULES = ["forward", "single", "multi", "rules"]', source)
        self.assertIn('storage.setItem', source)
        self.assertIn('drafts:', source)
        self.assertIn('paneWidths:', source)
        self.assertIn('shared:', source)

    def test_production_styles_include_narrow_screen_drawers(self):
        css = (self.web / "style.css").read_text(encoding="utf-8")
        self.assertIn('@media (max-width: 760px)', css)
        self.assertIn('.mobile-pane-tabs', css)
        self.assertIn('.workspace[data-mobile-pane=', css)

    def test_navigation_group_counts_and_five_status_dots(self):
        app = (self.web / "app.js").read_text(encoding="utf-8")
        html = (self.web / "index.html").read_text(encoding="utf-8")
        css = (self.web / "style.css").read_text(encoding="utf-8")
        # 组头展示 可见/总数 与相关计数，输入与输出两侧一致
        self.assertEqual(app.count("/${total} 项 · ${relevantCount} 相关"), 2)
        # 五种状态点并列：已选、已修改、约束、已有结果、异常
        for cls in ("selected", "edited", "constraint", "result", "error"):
            self.assertIn(f".dot.{cls}", css)
            self.assertIn(f'tag {cls}', html)
        # 输出导航同样渲染状态点
        self.assertIn("${outputStateDots(item)}", app)
        # 自动展开覆盖管理员默认项与活跃反向约束
        self.assertIn("display_defaults?.inputs?.includes", app)
        self.assertIn("display_defaults?.outputs?.includes", app)
        self.assertIn("activeReverseConstraints()", app)
        # 已有结果基于真实计算数据；异常覆盖规则错误与未命中约束
        self.assertIn("edited_values?.[item.id]", app)
        self.assertIn('RULE_ERROR_STATUSES = ["rejected", "unsupported"]', app)
        self.assertIn("x.hit === false", app)

    def test_navigation_collapsible_groups_and_star_pinning(self):
        app = (self.web / "app.js").read_text(encoding="utf-8")
        # relevant 仅作未手动操作时的默认展开，分组恒可折叠
        self.assertIn("navGroupOpen(", app)
        self.assertIn("stored === undefined", app)
        # 折叠时星标项留在外面
        self.assertIn("visibleItems", app)
        # 折叠按钮取反有效状态而非简单布尔翻转
        self.assertIn("openGroups[key] = !navGroupOpen(", app)

    def test_card_grid_adaptive_page_size(self):
        app = (self.web / "app.js").read_text(encoding="utf-8")
        css = (self.web / "style.css").read_text(encoding="utf-8")
        # 每页卡片数由可用宽/高动态计算，不再硬编码 6
        self.assertIn("function cardsPerPage()", app)
        self.assertNotIn("draft.page * 6", app)
        self.assertIn("ResizeObserver", app)
        # 网格列随宽度自适应
        self.assertIn("auto-fill", css)

    def test_output_pane_excel_hierarchy_and_collapsible_results(self):
        app = (self.web / "app.js").read_text(encoding="utf-8")
        html = (self.web / "index.html").read_text(encoding="utf-8")
        css = (self.web / "style.css").read_text(encoding="utf-8")
        state_js = (self.web / "workbench-state.js").read_text(encoding="utf-8")
        # 常驻摘要小卡片删除，#cards 仅承载瞬态结果且空时隐藏
        self.assertNotIn("renderCards(state.data.core_results)", app)
        self.assertIn('id="cards" class="cards" hidden', html)
        self.assertIn("function setCards(", app)
        # 结果表按 Excel 层次分节、节标题可折叠
        self.assertIn("section-head", app)
        self.assertIn("data-result-section", app)
        # 行内不再跟"未知 · 财务结果"式尾巴
        self.assertNotIn('${row.unit || ""} · ${row.group || ""}', app)
        # 指标行星标与分节折叠状态持久化
        self.assertIn("resultFavorites", app)
        self.assertIn("resultFavorites", state_js)
        self.assertIn("resultSections", state_js)

    def test_historical_template_read_only_switch_ui_hooks(self):
        app = (self.web / "app.js").read_text(encoding="utf-8")
        html = (self.web / "index.html").read_text(encoding="utf-8")
        css = (self.web / "style.css").read_text(encoding="utf-8")
        self.assertIn('id="templateSwitch"', html)
        self.assertIn('id="templateMode"', html)
        self.assertIn("renderTemplateSwitch", app)
        self.assertIn("historical_read_only", app)
        self.assertIn("isReadOnly()", app)
        self.assertIn("启用原因", app)
        self.assertIn(".badge.readonly", css)

    def test_card_inline_sliders_and_inline_reverse_config(self):
        app = (self.web / "app.js").read_text(encoding="utf-8")
        html = (self.web / "index.html").read_text(encoding="utf-8")
        css = (self.web / "style.css").read_text(encoding="utf-8")
        # 正向卡卡面内嵌五年滑杆：当前值/基准/差异，拖动走既有触发路径
        self.assertIn("data-card-slide", app)
        self.assertIn("year-tracks", app)
        self.assertIn("track-base", app)
        self.assertIn("applyYearValue", app)
        self.assertIn("scheduleAutomaticCalculation()", app)
        # 横/纵滑杆布局切换作用于卡片网格
        self.assertIn("card-grid layout-", app)
        self.assertIn(".card-grid.layout-horizontal", css)
        self.assertIn(".card-grid.layout-vertical", css)
        self.assertIn("vertical-range", css)
        # 变量卡卡面内联配置：上下限、范围滑杆、优先级、允许求解器搜索
        self.assertIn("data-v2c", app)
        self.assertIn("允许求解器搜索", app)
        self.assertIn("singleVariable", app)
        # 约束卡卡面内联配置：年份/关系/目标值/范围滑杆/软硬切换
        self.assertIn("data-cc", app)
        self.assertIn("切换为软目标", app)
        # 恢复基准保留；编辑器降级为高级入口
        self.assertIn("data-reset-card", app)
        self.assertIn('id="closeEditor"', html)


def rule(name="贷款利率", *, status="confirmed", linkage="independent", pending=False, allowed=(0, 10)):
    return {
        "rule_id": f"rule-{name}", "display_name": name, "indicator_group": "价格假设",
        "display_unit": "%", "adjustment_mode": "basis_point", "minimum_step": 0.01,
        "allowed_range": list(allowed), "linkage_strategy": linkage,
        "configuration_pending": pending, "confirmation_status": status, "rule_version": 2,
        "confirmed_source_cells": [{"year": str(year), "sheet": "信贷业务", "cell": f"C{year - 2000}"} for year in YEARS],
        "candidate_source_cells": [], "formula_dependency_chain": {}, "confidence": "high",
        "template_fingerprint": "fingerprint",
    }


class FakeTemplates:
    def __init__(self):
        self.catalog = [
            {"row": 6, "display_name": "归母净利润", "group": "财务结果", "unit": "亿元", "classification": "output", "cell_address": "B6", "year_values": {str(y): 100 + y - 2026 for y in YEARS}},
            {"row": 107, "display_name": "贷款利率", "group": "价格假设", "unit": "%", "classification": "input", "cell_address": "B107", "year_values": {str(y): 4 + (y - 2026) * .1 for y in YEARS}},
            {"row": 108, "display_name": "存款利率", "group": "价格假设", "unit": "%", "classification": "input", "cell_address": "B108", "year_values": {str(y): 2 for y in YEARS}},
        ]

    def list_template_versions(self):
        return [{"template_version": 1, "template_version_id": 1, "template_fingerprint": "fingerprint", "storage_id": "template.xlsx", "indicator_catalog": self.catalog, "worksheet": {"name": "汇总展示表", "index": 2}}]

    def get_indicator_catalog(self, template_version_id):
        return self.list_template_versions()[0]

    def regenerate_catalog(self, template_version_id, engine):
        return self.get_indicator_catalog(template_version_id)


class TwoVersionTemplates(FakeTemplates):
    def list_template_versions(self):
        return [
            {"template_version": 1, "template_version_id": 1, "template_fingerprint": "0716-fingerprint", "storage_id": "t0716.xlsx", "indicator_catalog": self.catalog, "worksheet": {"name": "汇总展示表", "index": 2}},
            {"template_version": 2, "template_version_id": 2, "template_fingerprint": "0717-fingerprint", "storage_id": "t0717.xlsx", "indicator_catalog": self.catalog, "worksheet": {"name": "汇总展示表", "index": 2}},
        ]

    def get_indicator_catalog(self, template_version_id):
        return next((item for item in self.list_template_versions() if item["template_version_id"] == template_version_id), None)


class FakeRules:
    def __init__(self, rules): self.rules = rules; self.active = 1
    def list_rules(self, template_version_id): return self.rules
    def get_rule(self, rule_id, include_snapshot=True): return next(item for item in self.rules if item["rule_id"] == rule_id)
    def confirm_and_configure(self, rule_id, *, expected_version, selected_sources, configuration, actor):
        current = next(item for item in self.rules if item["rule_id"] == rule_id)
        updated = {**current, **configuration, "rule_id": rule_id + "-confirmed", "rule_version": current["rule_version"] + 1, "confirmation_status": "confirmed", "configuration_pending": False, "confirmed_source_cells": selected_sources, "actor": actor}
        self.rules.append(updated); return updated
    def edit_rule(self, rule_id, *, expected_version, configuration, actor):
        current = next(item for item in self.rules if item["rule_id"] == rule_id)
        updated = {**current, **configuration, "rule_id": rule_id + "-edited", "rule_version": current["rule_version"] + 1, "actor": actor}
        self.rules.append(updated); return updated
    def reject_rule(self, rule_id, *, expected_version, reason, actor):
        current = next(item for item in self.rules if item["rule_id"] == rule_id)
        updated = {**current, "rule_id": rule_id + "-rejected", "rule_version": current["rule_version"] + 1, "confirmation_status": "rejected", "rejection_reason": reason, "actor": actor}
        self.rules.append(updated); return updated
    def rule_set_status(self, template_version_id):
        latest = {}
        for item in self.rules:
            latest[item.get("logical_rule_id", item["display_name"])] = item
        unresolved = [item for item in latest.values() if item["confirmation_status"] != "confirmed" or item.get("configuration_pending")]
        return {"template_version_id": template_version_id, "active": self.active == template_version_id, "complete": not unresolved, "unresolved_rules": unresolved}
    def activate_rule_set(self, template_version_id, template_fingerprint, *, actor):
        status = self.rule_set_status(template_version_id)
        if not status["complete"]: raise ValueError("规则集存在未解决规则")
        self.active = template_version_id; return {**status, "active": True, "actor": actor, "template_fingerprint": template_fingerprint}
    def deactivate_rule_set(self, template_version_id, *, actor):
        self.active = None; return {"template_version_id": template_version_id, "active": False, "actor": actor}
    def list_audit_logs(self, **kwargs): return []
    def get_rule_history(self, logical_rule_id): return [item for item in self.rules if item.get("logical_rule_id") == logical_rule_id]
    def get_rule_review(self, rule_id, chain_limit=20): return next(item for item in self.rules if item["rule_id"] == rule_id)
    def get_rule_history_summaries(self, logical_rule_id): return self.get_rule_history(logical_rule_id)
    def get_active_publication_rules(self, template_version_id): return self.rules if self.active == template_version_id else []
    def get_active_publication(self, template_version_id): return {"publication_id": "publication-1", "active": True} if self.active == template_version_id else None


class WorkbenchTests(unittest.TestCase):
    def service(self, rules=None, engine=None):
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        path = Path(directory.name) / "template.xlsx"
        path.write_bytes(b"template")
        return WorkbenchService(FakeTemplates(), FakeRules(rules or [rule(), rule("存款利率")]), lambda: engine or InMemoryWorkbookEngine(), Path(directory.name))

    def test_initialization_returns_template_catalog_rules_and_baseline(self):
        data = self.service().initialize()
        self.assertEqual(data["template"]["version"], 1)
        self.assertEqual(len(data["parameters"]), 2)
        self.assertEqual(data["baseline_results"]["归母净利润"]["2026"], 100)
        self.assertTrue(data["template"]["editable"])
        self.assertEqual(data["rule_set"]["publication_id"], "publication-1")
        self.assertEqual(data["scenario_draft"]["rule_publication_id"], "publication-1")

    def test_initialization_exposes_workbook_c_to_j_result_contract(self):
        data = self.service().initialize()
        row = next(item for item in data["result_rows"] if item["name"] == "归母净利润")
        self.assertEqual(row["values"]["2026"], 100)
        self.assertIn("2025", row["values"])
        self.assertIn("five_year_change", row["values"])
        self.assertIn("cagr", row["values"])

    def test_result_rows_carry_baseline_values_and_rule_precision(self):
        data = self.service().initialize()
        row = next(item for item in data["result_rows"] if item["name"] == "归母净利润")
        self.assertEqual(row["baseline_values"], row["values"])
        self.assertIsNone(row["precision"])
        self.assertEqual(WorkbenchService._display_precision({"minimum_step": 1}), 0)
        self.assertEqual(WorkbenchService._display_precision({"minimum_step": 0.01}), 2)
        self.assertEqual(WorkbenchService._display_precision({"minimum_step": 0.5}), 1)
        self.assertIsNone(WorkbenchService._display_precision({"minimum_step": None}))
        self.assertIsNone(WorkbenchService._display_precision({}))

    def test_calculation_result_rows_keep_pre_calculation_baseline(self):
        class ShiftingEngine(InMemoryWorkbookEngine):
            def read_summary(self, stage="summary_read"):
                base = {str(year): 100.0 for year in range(2026, 2031)}
                if stage == "baseline_summary_read":
                    return {"归母净利润": dict(base), "利润": dict(base)}
                return {"归母净利润": {year: value + 10 for year, value in base.items()}, "利润": dict(base)}

        result = self.service(engine=ShiftingEngine()).calculate(1, [{"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107", "values": {"2026": 4.2}}])
        self.assertEqual(result["trust"]["status"], "valid")
        row = next(item for item in result["result_rows"] if item["name"] == "归母净利润")
        self.assertEqual(row["values"]["2026"], 110)
        self.assertEqual(row["baseline_values"]["2026"], 100)

    def test_result_rows_interleave_section_headers_in_sheet_order(self):
        class SectionTemplates(FakeTemplates):
            def list_template_versions(self):
                return [{
                    **super().list_template_versions()[0],
                    "sections": [
                        {"row": 4, "title": "一、盈利结果", "level": 1},
                        {"row": 5, "title": "（一）并表口径", "level": 2},
                    ],
                }]

        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        service = WorkbenchService(SectionTemplates(), FakeRules([rule(), rule("存款利率")]), InMemoryWorkbookEngine, Path(directory.name))
        rows = service.initialize()["result_rows"]
        self.assertEqual([row["kind"] for row in rows], ["header", "header", "row"])
        self.assertEqual(rows[0]["title"], "一、盈利结果")
        self.assertEqual(rows[1]["title"], "（一）并表口径")
        self.assertEqual(rows[2]["name"], "归母净利润")
        self.assertIn("values", rows[2])

    def test_section_level_recognizes_numbered_and_parenthesized_titles(self):
        from forecast_engine import section_level
        self.assertEqual(section_level("一、盈利结果"), 1)
        self.assertEqual(section_level("（一）并表口径"), 2)
        self.assertEqual(section_level("（四）并表资本端变化"), 2)
        self.assertIsNone(section_level("并表RWA增量"))
        self.assertIsNone(section_level("归母净利润"))
        self.assertIsNone(section_level("资产端"))

    def test_capital_assumptions_group_classifies_as_input(self):
        from template_catalog import TemplateImportService
        item = {"row": 191, "display_name": "分红率", "group": "资本假设"}
        self.assertEqual(TemplateImportService._classify(item, {})["classification"], "input")


    def test_initialization_blocks_historical_template_workspace(self):
        service = WorkbenchService(FakeTemplates(), FakeRules([rule(), rule("存款利率")]), InMemoryWorkbookEngine, Path("."), "0717-fingerprint")
        with self.assertRaisesRegex(RuntimeError, "0717"):
            service.initialize()

    def test_initialization_without_publication_is_read_only(self):
        service = self.service()
        service.rules.active = None
        data = service.initialize()
        self.assertFalse(data["template"]["editable"])
        self.assertFalse(data["rule_set"]["active"])
        self.assertTrue(all(item["rule"] is None for item in data["parameters"]))
        self.assertIn("尚未发布", data["trust"]["reason"])

    def test_rule_admin_defaults_to_configured_activity_template(self):
        service = WorkbenchService(FakeTemplates(), FakeRules([rule(), rule("存款利率")]), InMemoryWorkbookEngine, Path("."), "fingerprint")
        self.assertEqual(service.rule_admin_bootstrap()["template"]["fingerprint"], "fingerprint")

    def test_rescan_is_idempotent_without_force(self):
        service = self.service()
        service.rules.list_latest_rule_summaries = lambda template_version_id: [{"logical_rule_id": "existing"}]
        result = service.rescan_template(1, actor="admin")
        self.assertTrue(result["skipped"])
        self.assertEqual(result["created_versions"], 0)

    def test_parameter_filters_cover_group_search_status_favorite_and_adjusted(self):
        items = [{"name": "贷款利率", "group": "价格假设", "rule_status": "confirmed", "favorite": True, "adjusted": True}, {"name": "存款规模", "group": "规模假设", "rule_status": "pending_confirmation", "favorite": False, "adjusted": False}]
        self.assertEqual([x["name"] for x in filter_parameters(items, search="贷款", group="价格假设", favorites=True, adjusted=True)], ["贷款利率"])
        self.assertEqual([x["name"] for x in filter_parameters(items, pending=True)], ["存款规模"])

    def test_linkage_strategies(self):
        baseline = {str(y): 100 + 10 * (y - 2026) for y in YEARS}
        self.assertEqual(apply_linkage(baseline, 2027, 135, "independent")["2026"], 100)
        self.assertEqual(apply_linkage(baseline, 2027, 135, "same_delta")["2030"], 165)
        self.assertEqual(set(apply_linkage(baseline, 2027, 135, "same_value").values()), {135})
        self.assertEqual(apply_linkage(baseline, 2027, 132, "baseline_ratio")["2030"], 168)

    def test_range_and_pending_rules_are_rejected(self):
        service = self.service([rule(allowed=(0, 5)), rule("存款利率", status="pending_confirmation")])
        with self.assertRaisesRegex(ValueError, "允许范围"):
            service.calculate(1, [{"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107", "values": {"2026": 6}}])
        result = service.calculate(1, [{"rule_id": "rule-存款利率", "indicator_id": "价格假设|存款利率|108", "values": {"2026": 2.1}}])
        self.assertEqual(result["trust"]["status"], "pending_rule_confirmation")

    def test_multiple_inputs_ignore_client_source_cells_and_return_details(self):
        engine = InMemoryWorkbookEngine()
        result = self.service(engine=engine).calculate(1, [
            {"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107", "values": {"2026": 4.2}, "source_cells": [{"sheet": "恶意", "cell": "A1"}]},
            {"rule_id": "rule-存款利率", "indicator_id": "价格假设|存款利率|108", "values": {"2027": 2.2}},
        ])
        self.assertEqual(result["trust"]["status"], "valid")
        self.assertEqual(len(result["calculation_details"]["submitted_adjustments"]), 2)
        self.assertNotIn("恶意", str(result["calculation_details"]["written_source_cells"]))
        self.assertEqual(result["calculation_details"]["rule_publication_id"], "publication-1")
        self.assertEqual(result["scenario_draft"]["validation_state"], "valid")
        self.assertEqual(result["scenario_draft"]["input_adjustments"]["价格假设|贷款利率|107"]["2026"], 4.2)
        self.assertIn("stage_timings", result["calculation_details"])

    def test_calculation_accepts_published_dict_source_mapping(self):
        published = rule()
        published["confirmed_source_cells"] = {str(year): {"sheet": "信贷业务", "cell": f"C{year - 2000}"} for year in YEARS}
        result = self.service([published], engine=InMemoryWorkbookEngine()).calculate(1, [{"rule_id": published["rule_id"], "indicator_id": "价格假设|贷款利率|107", "values": {"2026": 4.2}}])
        self.assertEqual(result["trust"]["status"], "valid")

    def test_missing_core_metric_is_not_fabricated(self):
        result = self.service().calculate(1, [])
        self.assertEqual([card["name"] for card in result["core_results"]], ["利润"])
        self.assertNotIn("LCR", str(result["core_results"]))

    def test_failed_calculation_preserves_edits(self):
        result = self.service(engine=InMemoryWorkbookEngine(fails=True)).calculate(1, [{"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107", "values": {"2026": 4.2}}])
        self.assertEqual(result["trust"]["status"], "calculation_failed")
        self.assertEqual(result["edited_values"]["价格假设|贷款利率|107"]["2026"], 4.2)

    def test_rule_admin_lists_latest_rules_and_status_counts(self):
        pending = rule(status="pending_confirmation", pending=True)
        pending["candidate_source_cells"] = [{"sheet": "参数", "year_cells": {str(year): f"{column}1" for year, column in zip(YEARS, "CDEFG")}, "reason": "formula_chain_ends_at_constant"}]
        data = self.service([pending, rule("存款利率")]).rule_admin(1, status="pending_confirmation")
        self.assertEqual(data["counts"]["pending_confirmation"], 1)
        self.assertEqual([item["display_name"] for item in data["rules"]], ["贷款利率"])
        self.assertEqual(data["rules"][0]["candidate_source_cells"][0]["year_cells"]["2026"], "C1")

    def test_admin_can_confirm_sources_and_adjustment_configuration(self):
        pending = rule(status="pending_confirmation", pending=True)
        service = self.service([pending])
        updated = service.update_rule(1, pending["rule_id"], {
            "action": "confirm", "expected_version": 2,
            "selected_sources": {str(year): {"sheet": "信贷业务", "cell": f"C{year - 2000}"} for year in YEARS},
            "adjustment_mode": "basis_point", "minimum_step": 1,
            "allowed_range": [0, 10], "linkage_strategy": "same_delta",
        })
        self.assertEqual(updated["confirmation_status"], "confirmed")
        self.assertFalse(updated["configuration_pending"])
        self.assertEqual(updated["rule_version"], 3)
        self.assertEqual(len(updated["confirmed_source_cells"]), 5)

    def test_admin_can_edit_or_reject_rule_as_new_version(self):
        service = self.service()
        edited = service.update_rule(1, "rule-贷款利率", {"action": "edit", "expected_version": 2, "adjustment_mode": "percentage_point", "minimum_step": .01, "allowed_range": [0, 8], "linkage_strategy": "independent"})
        rejected = service.update_rule(1, edited["rule_id"], {"action": "reject", "expected_version": 3, "rejection_reason": "不适用"})
        self.assertEqual(edited["adjustment_mode"], "percentage_point")
        self.assertEqual(rejected["confirmation_status"], "rejected")

    def test_confirmation_requires_all_five_year_sources(self):
        service = self.service([rule(status="pending_confirmation", pending=True)])
        with self.assertRaisesRegex(ValueError, "五个年度"):
            service.update_rule(1, "rule-贷款利率", {"action": "confirm", "expected_version": 2, "selected_sources": {"2026": {"sheet": "参数", "cell": "A1"}}, "adjustment_mode": "basis_point", "minimum_step": 1, "allowed_range": [0, 10], "linkage_strategy": "independent"})

    def test_incomplete_rule_set_cannot_be_activated(self):
        service = self.service([rule(status="pending_confirmation", pending=True)])
        status = service.rule_set(1)
        self.assertFalse(status["complete"])
        with self.assertRaisesRegex(ValueError, "未解决"):
            service.activate_rule_set(1, actor="admin")

    def test_rule_set_detects_missing_catalog_rule(self):
        status = self.service([rule()]).rule_set(1)
        self.assertFalse(status["complete"])
        self.assertEqual(status["missing_indicators"][0]["display_name"], "存款利率")

    def test_calculation_requires_active_rule_set(self):
        service = self.service()
        service.rules.active = None
        result = service.calculate(1, [{"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107", "values": {"2026": 4.2}}])
        self.assertEqual(result["trust"]["status"], "pending_rule_confirmation")
        self.assertIn("活动规则集", result["trust"]["reason"])

    def test_complete_rule_set_can_be_activated_and_deactivated(self):
        service = self.service()
        active = service.activate_rule_set(1, actor="admin")
        self.assertTrue(active["active"])
        self.assertFalse(service.deactivate_rule_set(1, actor="admin")["active"])

    def test_rule_detail_includes_history_and_audit(self):
        current = rule(); current["logical_rule_id"] = "logical-1"
        detail = self.service([current]).rule_detail(1, current["rule_id"])
        self.assertEqual(detail["rule"]["display_name"], "贷款利率")
        self.assertEqual(len(detail["history"]), 1)
        self.assertIn("audit", detail)
        self.assertEqual(detail["version_diff"], {})


class HistoricalTemplateReadOnlyTests(unittest.TestCase):
    def service(self, fingerprint="0717-fingerprint"):
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        path = Path(directory.name) / "template.xlsx"
        path.write_bytes(b"template")
        return WorkbenchService(TwoVersionTemplates(), FakeRules([rule(), rule("存款利率")]), InMemoryWorkbookEngine, Path(directory.name), fingerprint)

    def test_activity_template_is_default_and_lists_switchable_versions(self):
        data = self.service().initialize()
        self.assertEqual(data["template"]["fingerprint"], "0717-fingerprint")
        self.assertTrue(data["template"]["activity"])
        self.assertFalse(data["template"]["read_only"])
        historical = next(item for item in data["templates"] if item["fingerprint"] == "0716-fingerprint")
        self.assertTrue(historical["read_only"])
        self.assertFalse(historical["activity"])
        activity = next(item for item in data["templates"] if item["activity"])
        self.assertEqual(activity["id"], data["template"]["id"])

    def test_historical_template_view_is_read_only(self):
        service = self.service()
        service.rules.active = 1  # 历史模板存在活动发布，规则仍只读追溯
        data = service.initialize(1)
        self.assertTrue(data["template"]["read_only"])
        self.assertFalse(data["template"]["activity"])
        self.assertFalse(data["template"]["editable"])
        self.assertEqual(data["trust"]["status"], "historical_read_only")
        self.assertIn("只读", data["trust"]["reason"])
        self.assertIsNone(data["scenario_draft"])
        self.assertTrue(data["rule_set"]["active"])
        self.assertEqual([item["rule_status"] for item in data["parameters"]], ["confirmed", "confirmed"])
        row = next(item for item in data["result_rows"] if item["name"] == "归母净利润")
        self.assertEqual(row["values"]["2026"], 100)

    def test_unknown_template_version_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "模板版本不存在"):
            self.service().initialize(99)

    def test_historical_view_cannot_start_new_calculations(self):
        service = self.service()
        with self.assertRaisesRegex(ValueError, "历史模板"):
            service.start_calculation(1, [{"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107", "values": {"2026": 4.2}}])
        with self.assertRaisesRegex(ValueError, "历史模板"):
            service.start_reverse_calculation(1, {"variable": {"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107"}, "constraints": []})
        with self.assertRaisesRegex(ValueError, "历史模板"):
            service.save_scenario({"name": "历史场景", "template_version_id": 1, "scenario_type": "custom", "input_adjustments": {}})

    def test_activity_view_stays_editable_with_publication(self):
        service = self.service()
        service.rules.active = 2
        data = service.initialize()
        self.assertTrue(data["template"]["editable"])
        self.assertIsNotNone(data["scenario_draft"])


class SlowEngine(InMemoryWorkbookEngine):
    def __init__(self, delay=0.05, **kwargs):
        super().__init__(**kwargs)
        self.delay = delay

    def copy_cycle_ranges(self):
        time.sleep(self.delay)
        super().copy_cycle_ranges()


class LinearEngine(InMemoryWorkbookEngine):
    def read_summary(self, stage="summary_read"):
        value = float(self.adjustments.get(2026, 4.0))
        return {"利润": {str(year): value * 25 for year in YEARS}}


class MultiLinearEngine(InMemoryWorkbookEngine):
    histories = []

    def __init__(self, delay=0):
        super().__init__()
        self.values = {}
        self.delay = delay

    def write_input(self, rule, adjustment):
        self.values[rule.indicator] = dict(adjustment)

    def read_summary(self, stage="summary_read"):
        if self.delay:
            time.sleep(self.delay)
        loan = float(self.values.get("贷款利率", {}).get(2026, 4))
        deposit = float(self.values.get("存款利率", {}).get(2026, 2))
        self.__class__.histories.append((loan, deposit))
        return {"利润": {str(year): loan * 20 + deposit * 10 for year in YEARS}}


class FakeWarmWorker:
    instances = []

    def __init__(self, fingerprint, timeout_seconds=60):
        self.worker_id = f"fake-{len(self.instances) + 1}"
        self.calls = 0
        self.cleaned = False
        self.shutdown_called = False
        self.__class__.instances.append(self)

    def health(self):
        return {"healthy": True, "worker_id": self.worker_id, "queue_depth": 0, "error": None}

    def calculate(self, request, **kwargs):
        self.calls += 1
        result = run_forward_calculation(InMemoryWorkbookEngine(), request, template_path=kwargs["template_path"], cancel_token=kwargs.get("cancel_token"), progress=kwargs.get("progress"))
        result.update({"engine_mode": "warm_com", "worker_id": self.worker_id, "queue_wait_ms": 2.5, "cancel_status": "not_requested"})
        return result

    def cleanup_orphan(self):
        self.cleaned = True
        return True

    def shutdown(self):
        self.shutdown_called = True


class FailingWarmWorker(FakeWarmWorker):
    def calculate(self, request, **kwargs):
        raise WarmExcelWorkerError("simulated warm crash")


class UnhealthyWarmWorker(FakeWarmWorker):
    def health(self):
        return {"healthy": False, "worker_id": self.worker_id, "queue_depth": 0, "error": "simulated startup failure"}


class StatefulWarmWorker(FakeWarmWorker):
    def calculate(self, request, **kwargs):
        result = run_forward_calculation(LinearEngine(), request, template_path=kwargs["template_path"])
        result.update({"engine_mode": "warm_com", "worker_id": self.worker_id, "queue_wait_ms": 0, "cancel_status": "not_requested"})
        return result


class WarmEngineModeTests(unittest.TestCase):
    def setUp(self):
        FakeWarmWorker.instances = []

    def service(self, worker_factory=FakeWarmWorker):
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        (Path(directory.name) / "template.xlsx").write_bytes(b"template")
        return WorkbenchService(
            FakeTemplates(), FakeRules([rule(), rule("存款利率")]), InMemoryWorkbookEngine,
            Path(directory.name), "fingerprint", warm_worker_factory=worker_factory,
        )

    @staticmethod
    def adjustment(value=4.2):
        return [{"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107", "values": {"2026": value}}]

    def test_mode_selection_reuses_worker_and_reports_metadata(self):
        service = self.service()
        first = service.calculate(1, self.adjustment(), engine_mode="warm_com")
        second = service.calculate(1, self.adjustment(4.3), engine_mode="warm_com")
        self.assertEqual(first["calculation_details"]["engine_mode"], "warm_com")
        self.assertEqual(first["calculation_details"]["worker_id"], second["calculation_details"]["worker_id"])
        self.assertEqual(first["calculation_details"]["queue_wait_ms"], 2.5)
        self.assertEqual(FakeWarmWorker.instances[0].calls, 2)

    def test_ten_distinct_requests_do_not_share_input_state(self):
        service = self.service(StatefulWarmWorker)
        outputs = []
        for index in range(10):
            result = service.calculate(1, self.adjustment(4 + index / 10), engine_mode="warm_com")
            outputs.append(result["scenario_draft"]["calculation_result_snapshot"]["利润"]["2026"])
        for actual, expected in zip(outputs, [100 + index * 2.5 for index in range(10)]):
            self.assertAlmostEqual(actual, expected)

    def test_warm_failure_rebuilds_and_falls_back_to_cold(self):
        service = self.service(FailingWarmWorker)
        result = service.calculate(1, self.adjustment(), engine_mode="warm_com")
        details = result["calculation_details"]
        self.assertEqual(details["engine_mode_requested"], "warm_com")
        self.assertEqual(details["engine_mode"], "cold_com")
        self.assertIn("simulated warm crash", details["fallback_reason"])
        self.assertTrue(FakeWarmWorker.instances[0].shutdown_called)
        service.warm_worker_factory = FakeWarmWorker
        rebuilt = service.calculate(1, self.adjustment(), engine_mode="warm_com")
        self.assertEqual(rebuilt["calculation_details"]["engine_mode"], "warm_com")

    def test_invalid_mode_and_orphan_cleanup_entry(self):
        service = self.service()
        with self.assertRaisesRegex(ValueError, "engine mode"):
            service.calculate(1, self.adjustment(), engine_mode="unknown")
        service.calculate(1, self.adjustment(), engine_mode="warm_com")
        self.assertTrue(service.cleanup_warm_worker())
        self.assertTrue(FakeWarmWorker.instances[0].cleaned)

    def test_warm_recheck_starts_worker_once_and_reports_health(self):
        service = self.service()
        first = service.warm_worker_recheck()
        second = service.warm_worker_recheck()
        self.assertTrue(first["healthy"])
        self.assertEqual(first["worker_id"], "fake-1")
        self.assertEqual(second["worker_id"], "fake-1")
        self.assertEqual(len(FakeWarmWorker.instances), 1)

    def test_warm_recheck_replaces_unhealthy_worker(self):
        service = self.service(UnhealthyWarmWorker)
        failed = service.warm_worker_recheck()
        self.assertFalse(failed["healthy"])
        self.assertEqual(failed["error"], "simulated startup failure")
        service.warm_worker_factory = FakeWarmWorker
        recovered = service.warm_worker_recheck()
        self.assertTrue(FakeWarmWorker.instances[0].shutdown_called)
        self.assertTrue(recovered["healthy"])
        self.assertEqual(recovered["worker_id"], FakeWarmWorker.instances[1].worker_id)

    def test_warm_recheck_without_fingerprint_reports_unavailable(self):
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        service = WorkbenchService(
            FakeTemplates(), FakeRules([rule(), rule("存款利率")]), InMemoryWorkbookEngine,
            Path(directory.name), None, warm_worker_factory=FakeWarmWorker,
        )
        result = service.warm_worker_recheck()
        self.assertFalse(result["healthy"])
        self.assertIn("SHA-256", result["error"])
        self.assertEqual(FakeWarmWorker.instances, [])


class AsyncCalculationTests(unittest.TestCase):
    TERMINAL = ("succeeded", "failed", "cancelled", "cycle_not_converged")

    def service(self, engine, fingerprint=None):
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        path = Path(directory.name) / "template.xlsx"
        path.write_bytes(b"template")
        return WorkbenchService(FakeTemplates(), FakeRules([rule(), rule("存款利率")]), lambda: engine, Path(directory.name), fingerprint)

    def wait_task(self, service, task_id, timeout=10):
        deadline = time.time() + timeout
        while time.time() < deadline:
            snapshot = service.get_calculation(task_id)
            if snapshot["status"] in self.TERMINAL:
                return snapshot
            time.sleep(0.02)
        self.fail("测算任务未在超时内结束")

    def adjustment(self):
        return [{"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107", "values": {"2026": 4.2}}]

    def test_async_task_success_returns_result_and_timings(self):
        service = self.service(InMemoryWorkbookEngine())
        created = service.start_calculation(1, self.adjustment())
        self.assertIn(created["status"], ("queued", "running", "succeeded"))
        snapshot = self.wait_task(service, created["task_id"])
        self.assertEqual(snapshot["status"], "succeeded")
        self.assertEqual(snapshot["result"]["trust"]["status"], "valid")
        self.assertEqual(snapshot["result"]["scenario_draft"]["validation_state"], "valid")
        self.assertIn("stage_timings", snapshot["result"]["calculation_details"])
        self.assertEqual(snapshot["iterations"], 1)
        self.assertIsNotNone(snapshot["started_at"])
        self.assertIsNotNone(snapshot["finished_at"])
        self.assertGreaterEqual(snapshot["elapsed_ms"], 0)

    def test_async_task_failure_keeps_error(self):
        service = self.service(InMemoryWorkbookEngine(fails=True))
        created = service.start_calculation(1, self.adjustment())
        snapshot = self.wait_task(service, created["task_id"])
        self.assertEqual(snapshot["status"], "failed")
        self.assertIsNotNone(snapshot["error"])

    def test_async_task_cycle_not_converged_is_distinct_status(self):
        service = self.service(InMemoryWorkbookEngine(differences=[1.0]))
        created = service.start_calculation(1, self.adjustment())
        snapshot = self.wait_task(service, created["task_id"])
        self.assertEqual(snapshot["status"], "cycle_not_converged")
        self.assertEqual(snapshot["result"]["trust"]["status"], "cycle_not_converged")

    def test_cancel_request_then_cancelled_without_valid_result(self):
        service = self.service(SlowEngine(differences=[1.0]))
        created = service.start_calculation(1, self.adjustment())
        deadline = time.time() + 5
        while service.get_calculation(created["task_id"])["status"] not in ("running", "cancel_requested") and time.time() < deadline:
            time.sleep(0.01)
        requested = service.cancel_calculation(created["task_id"])
        self.assertEqual(requested["status"], "cancel_requested")
        self.assertTrue(requested["cancel_requested"])
        snapshot = self.wait_task(service, created["task_id"])
        self.assertEqual(snapshot["status"], "cancelled")
        self.assertIsNone(snapshot["result"])
        self.assertNotEqual(snapshot.get("result", {}) and snapshot["result"]["trust"]["status"], "valid")

    def test_cancel_terminal_task_is_noop(self):
        service = self.service(InMemoryWorkbookEngine())
        created = service.start_calculation(1, self.adjustment())
        snapshot = self.wait_task(service, created["task_id"])
        self.assertEqual(snapshot["status"], "succeeded")
        self.assertEqual(service.cancel_calculation(created["task_id"])["status"], "succeeded")

    def test_start_calculation_rejects_historical_template(self):
        service = self.service(InMemoryWorkbookEngine(), fingerprint="0717-fingerprint")
        with self.assertRaisesRegex(ValueError, "历史模板"):
            service.start_calculation(1, self.adjustment())

    def test_unknown_task_raises(self):
        service = self.service(InMemoryWorkbookEngine())
        with self.assertRaisesRegex(ValueError, "不存在"):
            service.get_calculation("missing-task")


class ScenarioTests(unittest.TestCase):
    def service(self, engine=None, fingerprint=None):
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        path = Path(directory.name) / "template.xlsx"
        path.write_bytes(b"template")
        return WorkbenchService(FakeTemplates(), FakeRules([rule(), rule("存款利率")]), lambda: engine or InMemoryWorkbookEngine(), Path(directory.name), fingerprint)

    def draft(self, **overrides):
        payload = {
            "name": "测试场景", "scenario_type": "custom", "template_version_id": 1,
            "rule_publication_id": "publication-1",
            "input_adjustments": {"价格假设|贷款利率|107": {"2026": 4.2}},
            "calculation_result_snapshot": {"归母净利润": {"2026": 101.0}},
            "validation_state": "valid",
        }
        payload.update(overrides)
        return payload

    def wait_task(self, service, task_id, timeout=10):
        deadline = time.time() + timeout
        while time.time() < deadline:
            snapshot = service.get_calculation(task_id)
            if snapshot["status"] in ("succeeded", "failed", "cancelled", "cycle_not_converged"):
                return snapshot
            time.sleep(0.02)
        self.fail("重算任务未在超时内结束")

    def test_save_and_open_scenario(self):
        service = self.service()
        saved = service.save_scenario(self.draft())
        self.assertFalse(saved["read_only"])
        self.assertEqual(saved["scenario_type"], "custom")
        opened = service.get_scenario(saved["scenario_id"])
        self.assertEqual(opened["input_adjustments"]["价格假设|贷款利率|107"]["2026"], 4.2)
        self.assertEqual(opened["calculation_result_snapshot"]["归母净利润"]["2026"], 101.0)
        self.assertEqual(opened["validation_state"], "valid")
        self.assertEqual(opened["rule_publication_id"], "publication-1")
        listed = service.list_scenarios()
        self.assertEqual(len(listed["scenarios"]), 1)
        self.assertEqual(listed["scenarios"][0]["adjustment_count"], 1)
        self.assertIn("reverse_result", listed["scenario_types"])

    def test_save_validates_name_type_and_historical_template(self):
        service = self.service()
        with self.assertRaisesRegex(ValueError, "名称"):
            service.save_scenario(self.draft(name=" "))
        with self.assertRaisesRegex(ValueError, "场景类型"):
            service.save_scenario(self.draft(scenario_type="unknown"))
        historical = self.service(fingerprint="0717-fingerprint")
        with self.assertRaisesRegex(ValueError, "历史模板"):
            historical.save_scenario(self.draft())

    def test_copy_scenario_inherits_content(self):
        service = self.service()
        saved = service.save_scenario(self.draft())
        copied = service.copy_scenario(saved["scenario_id"], {})
        self.assertNotEqual(copied["scenario_id"], saved["scenario_id"])
        self.assertEqual(copied["name"], "测试场景 副本")
        self.assertEqual(copied["input_adjustments"], saved["input_adjustments"])
        operations = [row["operation"] for row in service.scenarios.list_audit()]
        self.assertEqual(operations, ["scenario_created", "scenario_copied"])

    def test_rename_scenario_writes_audit(self):
        service = self.service()
        saved = service.save_scenario(self.draft())
        renamed = service.rename_scenario(saved["scenario_id"], {"name": "新名字"})
        self.assertEqual(renamed["name"], "新名字")
        with self.assertRaisesRegex(ValueError, "名称"):
            service.rename_scenario(saved["scenario_id"], {"name": ""})
        operations = [row["operation"] for row in service.scenarios.list_audit(saved["scenario_id"])]
        self.assertIn("scenario_renamed", operations)

    def test_delete_scenario_keeps_audit(self):
        service = self.service()
        saved = service.save_scenario(self.draft())
        service.delete_scenario(saved["scenario_id"])
        with self.assertRaisesRegex(ValueError, "不存在"):
            service.get_scenario(saved["scenario_id"])
        operations = [row["operation"] for row in service.scenarios.list_audit(saved["scenario_id"])]
        self.assertEqual(operations, ["scenario_created", "scenario_deleted"])

    def test_recalculate_scenario_updates_result(self):
        service = self.service()
        saved = service.save_scenario(self.draft(calculation_result_snapshot=None, validation_state=None))
        task = service.recalculate_scenario(saved["scenario_id"])
        snapshot = self.wait_task(service, task["task_id"])
        self.assertEqual(snapshot["status"], "succeeded")
        deadline = time.time() + 5
        while time.time() < deadline:
            updated = service.get_scenario(saved["scenario_id"])
            if updated["validation_state"] == "valid":
                break
            time.sleep(0.02)
        self.assertEqual(updated["validation_state"], "valid")
        self.assertIn("利润", updated["calculation_result_snapshot"])
        operations = [row["operation"] for row in service.scenarios.list_audit(saved["scenario_id"])]
        self.assertIn("scenario_recalculate_started", operations)
        self.assertIn("scenario_recalculated", operations)

    def test_recalculate_requires_active_rule_set(self):
        service = self.service()
        service.rules.active = None
        saved = service.save_scenario(self.draft())
        with self.assertRaisesRegex(ValueError, "活动规则集"):
            service.recalculate_scenario(saved["scenario_id"])

    def test_historical_scenario_is_read_only(self):
        service = self.service(fingerprint="0717-fingerprint")
        record = service.scenarios.create(
            name="0716 旧场景", scenario_type="custom", template_version_id=99,
            template_fingerprint="old-0716", rule_publication_id="old-publication",
            input_adjustments={"a": {"2026": 1}}, calculation_result_snapshot=None, validation_state="valid",
        )
        view = service.get_scenario(record["scenario_id"])
        self.assertTrue(view["read_only"])
        for action in (
            lambda: service.rename_scenario(record["scenario_id"], {"name": "x"}),
            lambda: service.delete_scenario(record["scenario_id"]),
            lambda: service.recalculate_scenario(record["scenario_id"]),
        ):
            with self.assertRaisesRegex(ValueError, "只读"):
                action()
        copied = service.copy_scenario(record["scenario_id"], {})
        self.assertTrue(copied["read_only"])


class ComparisonTests(ScenarioTests):
    def save(self, service, name, *, value=101.0, validation_state="valid", snapshot=True):
        return service.save_scenario(self.draft(
            name=name,
            calculation_result_snapshot={"归母净利润": {str(year): value + year - 2026 for year in YEARS}} if snapshot else None,
            validation_state=validation_state,
        ))

    def compare(self, service, scenarios, **overrides):
        request = {"scenario_ids": [item["scenario_id"] for item in scenarios], "baseline_scenario_id": scenarios[0]["scenario_id"], "comparison_id": "comparison-test"}
        request.update(overrides)
        return service.compare_scenarios(request)

    def test_requires_multiple_saved_scenarios(self):
        service = self.service()
        saved = self.save(service, "A")
        with self.assertRaisesRegex(ValueError, "至少选择两个"):
            service.start_comparison({"scenario_ids": [saved["scenario_id"]]})

    def test_valid_snapshots_are_used_without_recalculation_and_differences_are_computed(self):
        service = self.service(engine=InMemoryWorkbookEngine(fails=True))
        baseline = self.save(service, "基准", value=100)
        alternative = self.save(service, "方案", value=110)
        result = self.compare(service, [baseline, alternative])
        self.assertEqual([row["source"] for row in result["scenarios"]], ["snapshot", "snapshot"])
        self.assertEqual(result["summary"], {"total": 2, "valid": 2, "failed": 0})
        self.assertEqual(result["core_results"][0]["scenarios"][1]["differences"]["2026"], 10)
        self.assertEqual(result["details"][0]["scenarios"][1]["differences"]["2030"], 10)

    def test_missing_or_invalid_snapshot_triggers_recalculation(self):
        service = self.service()
        baseline = self.save(service, "基准", value=100)
        missing = self.save(service, "待重算", snapshot=False, validation_state=None)
        result = self.compare(service, [baseline, missing])
        recalculated = result["scenarios"][1]
        self.assertEqual(recalculated["source"], "recalculated")
        self.assertEqual(recalculated["status"], "succeeded")
        self.assertIsNotNone(recalculated["calculation_details"])

    def test_force_refresh_recalculates_valid_snapshot(self):
        service = self.service()
        baseline = self.save(service, "基准", value=100)
        alternative = self.save(service, "方案", value=110)
        result = self.compare(service, [baseline, alternative], force_refresh=True)
        self.assertTrue(all(row["source"] == "recalculated" for row in result["scenarios"]))
        details = result["calculation_details"]
        self.assertEqual(details["engine_mode_requested"], None)
        self.assertEqual(details["engine_mode"], "cold_com")
        self.assertEqual(set(details["stage_timings"]), {"scenario_1", "scenario_2"})

    def test_single_scenario_failure_is_isolated(self):
        service = self.service(engine=InMemoryWorkbookEngine(fails=True))
        baseline = self.save(service, "基准", value=100)
        failing = self.save(service, "失败场景", snapshot=False, validation_state=None)
        result = self.compare(service, [baseline, failing])
        self.assertEqual(result["summary"], {"total": 2, "valid": 1, "failed": 1})
        self.assertEqual(result["scenarios"][1]["status"], "calculation_failed")
        self.assertIsNone(result["scenarios"][1]["calculation_result_snapshot"])

    def test_historical_read_only_uses_valid_snapshot_but_does_not_recalculate(self):
        service = self.service(fingerprint="fingerprint")
        baseline = self.save(service, "基准", value=100)
        historical = service.scenarios.create(
            name="0716", scenario_type="custom", template_version_id=99, template_fingerprint="old",
            rule_publication_id="old-publication", input_adjustments={}, calculation_result_snapshot=None, validation_state=None,
        )
        result = self.compare(service, [baseline, historical])
        self.assertTrue(result["scenarios"][1]["read_only"])
        self.assertIn("历史只读", result["scenarios"][1]["failure_reason"])

    def test_async_comparison_can_be_cancelled_and_writes_audit(self):
        service = self.service(engine=SlowEngine(delay=.1))
        baseline = self.save(service, "基准", snapshot=False, validation_state=None)
        alternative = self.save(service, "方案", snapshot=False, validation_state=None)
        created = service.start_comparison({"scenario_ids": [baseline["scenario_id"], alternative["scenario_id"]], "force_refresh": True})
        service.cancel_calculation(created["task_id"])
        task = self.wait_task(service, created["task_id"])
        self.assertEqual(task["status"], "cancelled")
        operations = [row["operation"] for row in service.scenarios.list_audit()]
        self.assertIn("comparison_started", operations)
        self.assertIn("comparison_failed", operations)


class ReverseCalculationTests(unittest.TestCase):
    TERMINAL = ("succeeded", "failed", "cancelled", "cycle_not_converged")

    def service(self, engine_factory=LinearEngine, rules=None):
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        (Path(directory.name) / "template.xlsx").write_bytes(b"template")
        return WorkbenchService(FakeTemplates(), FakeRules(rules or [rule(), rule("存款利率")]), engine_factory, Path(directory.name))

    def request(self, **overrides):
        payload = {
            "variable": {"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107", "year": "2026", "lower": 0, "upper": 10},
            "constraints": [{"indicator_name": "利润", "indicator_type": "output", "year": "2026", "kind": "target", "value": 100, "tolerance": 0.1, "hard": True, "enabled": True}],
            "max_evaluations": 25,
        }
        payload.update(overrides)
        return payload

    def wait(self, service, task_id):
        deadline = time.time() + 10
        while time.time() < deadline:
            task = service.get_calculation(task_id)
            if task["status"] in self.TERMINAL:
                return task
            time.sleep(.01)
        self.fail("反向测算超时")

    def test_feasible_solution_and_reverse_scenario(self):
        service = self.service()
        result = service.reverse_calculate(1, self.request())
        self.assertTrue(result["feasible"])
        self.assertAlmostEqual(result["variable"]["required_value"], 4.0, places=2)
        self.assertTrue(result["constraints"][0]["hit"])
        self.assertEqual(result["scenario_draft"]["scenario_type"], "reverse_result")
        saved = service.save_scenario({"name": "反向结果", **result["scenario_draft"]})
        self.assertEqual(saved["scenario_type"], "reverse_result")
        self.assertEqual(service.reverse_audit[0]["result_status"], "valid")
        self.assertEqual(service.scenarios.list_audit()[0]["operation"], "reverse_calculation")

    def test_no_feasible_hard_constraint(self):
        result = self.service().reverse_calculate(1, self.request(constraints=[{"indicator_name": "利润", "year": "2026", "kind": "min", "value": 300, "hard": True}]))
        self.assertFalse(result["feasible"])
        self.assertEqual(result["trust"]["status"], "reverse_no_feasible")
        self.assertIsNone(result["scenario_draft"])

    def test_variable_result_explains_enable_reason(self):
        result = self.service().reverse_calculate(1, self.request())
        self.assertTrue(result["feasible"])
        self.assertIn("唯一求解变量", result["variable"]["reason"])
        self.assertFalse(result["variable"]["hit_boundary"])

    def test_infeasible_search_marks_boundary_in_enable_reason(self):
        result = self.service().reverse_calculate(1, self.request(constraints=[{"indicator_name": "利润", "year": "2026", "kind": "min", "value": 300, "hard": True}]))
        self.assertFalse(result["feasible"])
        self.assertTrue(result["variable"]["hit_boundary"])
        self.assertIn("边界", result["variable"]["reason"])

    def test_soft_constraint_reports_deviation(self):
        result = self.service().reverse_calculate(1, self.request(constraints=[
            {"indicator_name": "利润", "year": "2026", "kind": "min", "value": 100, "hard": True},
            {"indicator_name": "利润", "year": "2026", "kind": "target", "value": 110, "hard": False},
        ]))
        self.assertTrue(result["feasible"])
        self.assertGreaterEqual(result["soft_deviation"], 0)
        self.assertEqual(len(result["constraints"]), 2)

    def test_unconfirmed_variable_is_rejected(self):
        with self.assertRaisesRegex(ValueError, "未确认"):
            self.service(rules=[rule(status="pending_confirmation"), rule("存款利率")]).reverse_calculate(1, self.request())

    def test_calculation_failure_is_explicit(self):
        with self.assertRaisesRegex(RuntimeError, "正向计算失败"):
            self.service(lambda: InMemoryWorkbookEngine(fails=True)).reverse_calculate(1, self.request())

    def test_async_progress_and_cancel(self):
        service = self.service(lambda: SlowEngine(delay=.05))
        created = service.start_reverse_calculation(1, self.request(max_evaluations=25))
        service.cancel_calculation(created["task_id"])
        task = self.wait(service, created["task_id"])
        self.assertEqual(task["status"], "cancelled")
        self.assertIsNone(task["result"])

    def test_single_variable_search_starts_at_configured_initial_value(self):
        seen = []
        result = search_single_variable(
            lower=0,
            upper=10,
            initial=7,
            max_evaluations=3,
            evaluate=lambda value, _index: seen.append(value) or {"hard_violation": 0, "soft_deviation": 0},
        )
        self.assertEqual(seen[0], 7)
        self.assertEqual(result["variable_value"], 7)


class ReverseCalculationV2Tests(unittest.TestCase):
    TERMINAL = ("succeeded", "failed", "cancelled", "cycle_not_converged")

    def setUp(self):
        MultiLinearEngine.histories = []

    def service(self, engine_factory=MultiLinearEngine, rules=None):
        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        (Path(directory.name) / "template.xlsx").write_bytes(b"template")
        return WorkbenchService(FakeTemplates(), FakeRules(rules or [rule(), rule("存款利率")]), engine_factory, Path(directory.name))

    def wait(self, service, task_id):
        deadline = time.time() + 10
        while time.time() < deadline:
            task = service.get_calculation(task_id)
            if task["status"] in self.TERMINAL:
                return task
            time.sleep(.01)
        self.fail("反向测算超时")

    def request_v2(self, **overrides):
        payload = {
            "variables": [
                {"rule_id": "rule-贷款利率", "indicator_id": "价格假设|贷款利率|107", "year": "2026", "priority": 1, "lower": 4, "upper": 5, "step": 1, "linkage_strategy": "independent"},
                {"rule_id": "rule-存款利率", "indicator_id": "价格假设|存款利率|108", "year": "2026", "priority": 2, "lower": 2, "upper": 5, "step": 1, "linkage_strategy": "independent"},
            ],
            "constraints": [{"indicator_name": "利润", "indicator_type": "output", "year": "2026", "kind": "min", "value": 150, "hard": True, "enabled": True}],
            "max_evaluations": 10,
        }
        payload.update(overrides)
        return payload

    def test_priority_search_finds_multi_variable_solution_without_explosion(self):
        result = self.service(MultiLinearEngine).reverse_calculate(1, self.request_v2())
        self.assertTrue(result["feasible"])
        self.assertEqual([item["indicator_name"] for item in result["variables"]], ["贷款利率", "存款利率"])
        self.assertEqual([item["suggested_value"] for item in result["variables"]], [5, 5])
        self.assertEqual([item["indicator_id"] for item in result["variables"] if item["hit_boundary"]], ["价格假设|贷款利率|107", "价格假设|存款利率|108"])
        self.assertEqual([item["key"] for item in result["adjustment_path"]], ["价格假设|贷款利率|107", "价格假设|存款利率|108"])
        self.assertEqual(result["search_count"], 5)
        self.assertEqual(MultiLinearEngine.histories[1::2][:2], [(4, 2), (5, 2)])

    def test_same_priority_variables_can_be_combined_and_ranked_by_changed_count(self):
        variables = [
            {"key": "a", "priority": 1, "order": 0, "baseline": 0, "lower": 0, "upper": 1, "candidates": [0, 1]},
            {"key": "b", "priority": 1, "order": 1, "baseline": 0, "lower": 0, "upper": 1, "candidates": [0, 1]},
        ]

        def evaluate(values, _index):
            total = values["a"] + values["b"]
            return {"hard_violation": max(0, 2 - total), "soft_deviation": 0}

        result = search_priority_variables(variables=variables, evaluate=evaluate, max_evaluations=8)
        self.assertTrue(result["feasible"])
        self.assertEqual(result["variable_values"], {"a": 1, "b": 1})

    def test_no_solution_returns_closest_boundaries_and_reason(self):
        request = self.request_v2(constraints=[{"indicator_name": "利润", "year": "2026", "kind": "min", "value": 200, "hard": True}])
        result = self.service(MultiLinearEngine).reverse_calculate(1, request)
        self.assertFalse(result["feasible"])
        self.assertEqual(result["trust"]["status"], "reverse_no_feasible")
        self.assertTrue(result["no_feasible_reason"])
        self.assertEqual([item["suggested_value"] for item in result["variables"]], [5, 5])
        self.assertTrue(result["searched_ranges"])
        self.assertIsNone(result["scenario_draft"])

    def test_variables_explain_enable_reasons(self):
        result = self.service(MultiLinearEngine).reverse_calculate(1, self.request_v2())
        reasons = {item["indicator_name"]: item["reason"] for item in result["variables"]}
        self.assertIn("最高优先级", reasons["贷款利率"])
        self.assertIn("硬约束缺口", reasons["贷款利率"])
        self.assertIn("更高优先级", reasons["存款利率"])
        self.assertIn("边界", reasons["存款利率"])

    def test_unused_variable_reports_not_enabled_reason(self):
        request = self.request_v2(constraints=[{"indicator_name": "利润", "year": "2026", "kind": "min", "value": 120, "hard": True}])
        result = self.service(MultiLinearEngine).reverse_calculate(1, request)
        self.assertTrue(result["feasible"])
        deposit = next(item for item in result["variables"] if item["indicator_name"] == "存款利率")
        self.assertEqual(deposit["suggested_value"], 2)
        self.assertIn("未启用", deposit["reason"])

    def test_soft_constraint_deviation_and_reverse_result_save(self):
        request = self.request_v2(constraints=[
            {"indicator_name": "利润", "year": "2026", "kind": "min", "value": 150, "hard": True},
            {"indicator_name": "利润", "year": "2026", "kind": "target", "value": 160, "hard": False},
        ])
        service = self.service(MultiLinearEngine)
        result = service.reverse_calculate(1, request)
        self.assertEqual(result["soft_deviation"], 10)
        saved = service.save_scenario({"name": "v2 反向结果", **result["scenario_draft"]})
        self.assertEqual(saved["scenario_type"], "reverse_result")
        self.assertEqual(len(saved["input_adjustments"]), 2)

    def test_soft_only_target_is_searched_instead_of_stopping_at_baseline(self):
        request = self.request_v2(constraints=[{"indicator_name": "利润", "year": "2026", "kind": "target", "value": 150, "hard": False}])
        result = self.service(MultiLinearEngine).reverse_calculate(1, request)
        self.assertEqual(result["soft_deviation"], 0)
        self.assertGreater(result["search_count"], 1)

    def test_v2_async_cancel_and_budget_validation(self):
        service = self.service(lambda: MultiLinearEngine(delay=.03))
        created = service.start_reverse_calculation(1, self.request_v2(max_evaluations=10))
        service.cancel_calculation(created["task_id"])
        task = self.wait(service, created["task_id"])
        self.assertEqual(task["status"], "cancelled")
        with self.assertRaisesRegex(ValueError, "2–20"):
            self.service(MultiLinearEngine).reverse_calculate(1, self.request_v2(max_evaluations=21))


class ExportTests(ComparisonTests):
    def metadata(self, **overrides):
        value = {
            "template_version_id": 1, "template_fingerprint": "fingerprint",
            "rule_publication_id": "publication-1", "scenario_id": "scenario-1",
            "scenario_type": "custom", "calculation_time": "2026-07-17T12:00:00+00:00",
            "validation_state": "valid",
        }
        value.update(overrides)
        return value

    def workbook(self, service, exported):
        return load_workbook(service.get_export(exported["file_id"]), data_only=True)

    def test_exports_forward_scenario_with_complete_metadata(self):
        service = self.service()
        exported = service.export("scenario", {
            "metadata": self.metadata(),
            "scenario_draft": self.draft(),
            "details": [{"group": "财务结果", "name": "归母净利润", "values": {"2026": 101.0}}],
        })
        workbook = self.workbook(service, exported)
        self.assertEqual(workbook.sheetnames, ["Metadata", "Inputs", "Results", "Details"])
        metadata = dict(workbook["Metadata"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(metadata["template_fingerprint"], "fingerprint")
        self.assertEqual(metadata["validation_state"], "valid")
        self.assertEqual(workbook["Results"]["A2"].value, "归母净利润")
        self.assertTrue(Path(exported["path"]).is_file())

    def test_exports_reverse_result_details(self):
        service = self.service()
        reverse = {
            "variable": {"indicator_name": "贷款利率", "year": "2026", "required_value": 4.2, "adjustment": 0.2},
            "constraints": [{"indicator_name": "归母净利润", "year": "2026", "kind": "min", "hard": True, "value": 100, "actual": 101, "hit": True, "deviation": 0}],
            "feasible": True, "soft_deviation": 0, "search_count": 5,
        }
        exported = service.export("reverse", {**reverse, "metadata": self.metadata(scenario_type="reverse_result")})
        workbook = self.workbook(service, exported)
        self.assertEqual(workbook.sheetnames, ["Metadata", "Variable", "Results", "Constraints"])
        variable = dict(workbook["Variable"].iter_rows(min_row=2, values_only=True))
        results = dict(workbook["Results"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(variable["indicator_name"], "贷款利率")
        self.assertGreater(results["search_count"], 0)
        self.assertEqual(workbook["Constraints"]["G2"].value, True)

    def test_exports_v2_reverse_variables(self):
        service = self.service()
        reverse = {
            "variable": {"indicator_name": "贷款利率", "year": "2026", "required_value": 5, "adjustment": 1},
            "variables": [
                {"indicator_name": "贷款利率", "year": "2026", "priority": 1, "baseline_value": 4, "suggested_value": 5, "adjustment": 1, "lower": 4, "upper": 5, "hit_boundary": True, "linkage_strategy": "independent"},
                {"indicator_name": "存款利率", "year": "2026", "priority": 2, "baseline_value": 2, "suggested_value": 5, "adjustment": 3, "lower": 2, "upper": 5, "hit_boundary": True, "linkage_strategy": "same_value"},
            ],
            "constraints": [{"indicator_name": "利润", "year": "2026", "kind": "min", "hard": True, "value": 150, "actual": 150, "hit": True, "deviation": 0}],
            "feasible": True, "soft_deviation": 0, "search_count": 5,
        }
        exported = service.export("reverse", {**reverse, "metadata": self.metadata(scenario_type="reverse_result")})
        workbook = self.workbook(service, exported)
        self.assertEqual(workbook.sheetnames, ["Metadata", "Variables", "Results", "Constraints"])
        self.assertEqual(workbook["Variables"]["A3"].value, "存款利率")
        self.assertEqual(workbook["Variables"]["I2"].value, True)

    def test_exports_comparison_values_differences_and_states(self):
        service = self.service()
        baseline = self.save(service, "基准", value=100)
        alternative = self.save(service, "方案", value=110)
        comparison = self.compare(service, [baseline, alternative])
        exported = service.export("comparison", comparison)
        workbook = self.workbook(service, exported)
        self.assertEqual(workbook.sheetnames, ["Metadata", "Scenarios", "Comparison"])
        rows = list(workbook["Comparison"].iter_rows(min_row=2, values_only=True))
        self.assertTrue(any(row[3] == "方案" and row[6] == 10 and row[7] == "valid" for row in rows))
        metadata = dict(workbook["Metadata"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(metadata["scenario_type"], "comparison")
        self.assertEqual(metadata["scenario_id"], "comparison-test")

    def test_export_success_and_failure_are_audited(self):
        service = self.service()
        service.export("scenario", {"metadata": self.metadata(), "scenario_draft": self.draft()})
        with self.assertRaisesRegex(ValueError, "元数据不完整"):
            service.export("scenario", {"scenario_draft": {}})
        operations = [row["operation"] for row in service.scenarios.list_audit()]
        self.assertEqual(operations, ["export_succeeded", "export_failed"])


if __name__ == "__main__":
    unittest.main()
